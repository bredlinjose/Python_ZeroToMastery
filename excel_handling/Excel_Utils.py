import os

import openpyxl

excel_path = os.path.join(os.path.dirname(__file__), "sampledata.xlsx")


def get_row_count(sheet_name):
    workbook = openpyxl.open(excel_path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(sheet_name):
    workbook = openpyxl.open(excel_path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data_from_excel(sheet_name, row, column):
    workbook = openpyxl.open(excel_path)
    sheet = workbook[sheet_name]
    return sheet.cell(row, column).value


def write_data_into_excel(sheet_name, row, column, data):
    workbook = openpyxl.open(excel_path)
    sheet = workbook[sheet_name]
    sheet.cell(row, column).value = data
    workbook.save(excel_path)


def fetch_data_from_excel(sheet_name, row_header, column_header):
    workbook = openpyxl.open(excel_path)
    sheet = workbook[sheet_name]

    # Find the row and column indices
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_header in row:
            break
    for col_index, col in enumerate(sheet.iter_cols(values_only=True), start=1):
        if column_header in col:
            break

    # Check if the headers were found
    if row_index and col_index:
        cell_value = sheet.cell(row_index, col_index).value
        return cell_value
    else:
        return None

def get_row_excel_data(sheet_name, row, path=excel_path):
    data = []
    workbook = openpyxl.open(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_columns = sheet.max_column
    for i in range(row, total_rows + 1):
        for column in range(1, total_columns + 1):
            data.append(sheet.cell(i, column).value)
    return data


def fetch_data_by_row_header(sheet_name, row_header_value, path=excel_path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]

    # Find the row index based on the row header value
    row_index = None
    for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_header_value in row:
            row_index = row_num
            break
    if row_index is None:
        raise ValueError(f"Row header '{row_header_value}' not found in sheet '{sheet_name}'.")
    # Extract data from the specified row
    data = []
    total_rows = sheet.max_row
    total_columns = sheet.max_column
    for i in range(row_index, total_rows + 1):
        for column in range(1, total_columns + 1):
            data.append(sheet.cell(i, column).value)
        break
    return data

def get_row_number(row_header, sheet_name, path=excel_path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    row_index = None
    for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_header in row:
            row_index = row_num
            break
    print(row_index)

def get_column_number(column_header, sheet_name, path=excel_path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    col_index = None
    for col_index, col in enumerate(sheet.iter_cols(values_only=True), start=1):
        if column_header in col:
            break
    print(col_index)

def extract_whole_data_as_dicts(filename, sheet_name):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheet_name]

        # Extract column headers
        headers = [cell.value for cell in sheet[1]]

        # Extract data rows
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)

        return data

    except Exception as e:
        print(f"Error: {e}")
        return []

def extract_data_as_dicts(row_header, sheet_name, filename=excel_path):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheet_name]

        # Find the row index based on the row header value
        row_index = None
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_header in row:
                row_index = row_num
                break

        if row_index is None:
            raise ValueError(f"Row header '{row_header}' not found in sheet '{sheet_name}'.")

        # Extract column headers
        headers = [cell.value for cell in sheet[1]]

        # Extract data rows
        row_dict = ()
        for row in sheet.iter_rows(min_row=row_index, values_only=True):
            print(row)
            row_dict = dict(zip(headers, row))
            break
        return row_dict

    except Exception as e:
        print(f"Error: {e}")
        return ()

def fetch_data_as_dicts(sheet_name, filename=excel_path):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        row_dict = None
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
        return row_dict
    except Exception as e:
        print(f"Error: {e}")
        return None


def fetch_row_header_using_key(key, sheet_name, filename= excel_path):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]

    row_header = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[1] == key:
            row_header.append(row[0])

    return row_header


print(fetch_row_header_using_key("Yes", "Dummy"))

def fetch_modules_with_execution(key, sheet_name, module_col, execution_col, filename=excel_path):

    wb = openpyxl.load_workbook(filename)
    ws = wb[sheet_name]

    # Find the column indices based on header names
    module_col_index = None
    execution_col_index = None
    for col_index, cell in enumerate(ws[1]):
        if cell.value == module_col:
            module_col_index = col_index
        if cell.value == execution_col:
            execution_col_index = col_index

    if module_col_index is None or execution_col_index is None:
        raise ValueError("Header names not found in the spreadsheet.")

    module_names = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[execution_col_index] == "Yes":
            module_names.append(row[module_col_index])

    return module_names
